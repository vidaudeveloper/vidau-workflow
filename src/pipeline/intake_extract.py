"""从上传文件中提取文本或 Gemini 多模态 part。"""

from __future__ import annotations

import base64
import csv
import io
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from src.pipeline.intake_formats import (
    MAX_TEXT_EXCERPT_CHARS,
    VIDEO_URL_RE,
    IntakeFormat,
    classify_intake_filename,
)

_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _truncate(text: str, limit: int = MAX_TEXT_EXCERPT_CHARS) -> str:
    text = text.strip()
    if len(text) <= limit:
        return text
    return text[:limit] + f"\n…(已截断，原文 {len(text)} 字符)"


def _inline_bytes(data: bytes, mime: str) -> dict[str, Any]:
    return {
        "inline_data": {
            "mime_type": mime,
            "data": base64.b64encode(data).decode("ascii"),
        }
    }


def _extract_docx_text(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        xml = zf.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    paragraphs: list[str] = []
    for para in root.iter(f"{_W_NS}p"):
        runs = [node.text for node in para.iter(f"{_W_NS}t") if node.text]
        line = "".join(runs).strip()
        if line:
            paragraphs.append(line)
    return _truncate("\n".join(paragraphs))


def _extract_csv_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            raw = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raw = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(raw))
    rows = []
    for i, row in enumerate(reader):
        if i >= 400:
            rows.append("…(更多行已省略)")
            break
        rows.append("\t".join(cell.strip() for cell in row))
    return _truncate("CSV 表格内容:\n" + "\n".join(rows))


def _extract_xlsx_text(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("服务器未安装 openpyxl，无法解析 Excel") from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    chunks: list[str] = []
    for sheet_name in wb.sheetnames[:5]:
        ws = wb[sheet_name]
        chunks.append(f"## Sheet: {sheet_name}")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                chunks.append("\t".join(cells))
            row_count += 1
            if row_count >= 200:
                chunks.append("…(更多行已省略)")
                break
    wb.close()
    return _truncate("\n".join(chunks))


def _extract_plain_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return _truncate(data.decode(encoding))
        except UnicodeDecodeError:
            continue
    return _truncate(data.decode("utf-8", errors="replace"))


def extract_text_content(data: bytes, filename: str, fmt: IntakeFormat) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".docx":
        return _extract_docx_text(data)
    if ext == ".csv":
        return _extract_csv_text(data)
    if ext == ".xlsx":
        return _extract_xlsx_text(data)
    if ext in (".txt", ".md"):
        return _extract_plain_text(data)
    if fmt.category == "table":
        return _extract_plain_text(data)
    raise ValueError(f"无法从 {filename} 提取文本")


def build_multimodal_part(data: bytes, filename: str, fmt: IntakeFormat) -> dict[str, Any]:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return _inline_bytes(data, "application/pdf")
    if fmt.category == "reference_video":
        mime = {
            ".mp4": "video/mp4",
            ".webm": "video/webm",
            ".mov": "video/quicktime",
            ".m4v": "video/x-m4v",
        }.get(ext, "video/mp4")
        return _inline_bytes(data, mime)
    if fmt.category == "product_image":
        mime = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".webp": "image/webp",
            ".heic": "image/heic",
            ".heif": "image/heif",
        }.get(ext, "image/jpeg")
        return _inline_bytes(data, mime)
    raise ValueError(f"不支持的二进制素材: {filename}")


def _text_part(text: str) -> dict[str, Any]:
    return {"text": text}


def part_label_for_format(fmt: IntakeFormat, filename: str) -> dict[str, Any]:
    labels = {
        "product_doc": "产品介绍文档",
        "table": "规格/卖点表格",
        "reference_video": "对标参考视频",
        "product_image": "产品参考图",
        "text": "文本 Brief",
    }
    label = labels.get(fmt.category, fmt.category)
    return _text_part(f"--- {label}: {filename} ---")


def normalize_gemini_parts(parts: list[Any]) -> list[dict[str, Any]]:
    """Vertex / Gemini 要求 parts 均为 {text} 或 {inline_data} 对象，不能是裸字符串。"""
    out: list[dict[str, Any]] = []
    for part in parts:
        if isinstance(part, str):
            out.append(_text_part(part))
        elif isinstance(part, dict):
            out.append(part)
    return out


def validate_attachment_size(data: bytes, fmt: IntakeFormat, filename: str) -> None:
    if len(data) > fmt.max_bytes:
        mb = fmt.max_bytes // (1024 * 1024)
        raise ValueError(f"{filename} 超过 {mb}MB 限制")


def prepare_attachment(
    data: bytes, filename: str
) -> tuple[IntakeFormat, list[dict[str, Any]]]:
    """返回 (格式, Gemini user parts 片段)。"""
    fmt = classify_intake_filename(filename)
    if not fmt:
        raise ValueError(
            f"不支持的文件类型: {filename}。"
            "支持 PDF、Word、CSV、Excel、对标视频、产品图、TXT/MD"
        )
    if not data:
        raise ValueError(f"{filename} 为空")
    validate_attachment_size(data, fmt, filename)

    ext = Path(filename).suffix.lower()
    parts: list[dict[str, Any]] = [part_label_for_format(fmt, filename)]
    if fmt.category in ("table", "text") or ext == ".docx":
        text = extract_text_content(data, filename, fmt)
        parts.append(_text_part(text))
    else:
        parts.append(build_multimodal_part(data, filename, fmt))
    return fmt, parts


GENERIC_URL_RE = re.compile(r"https?://[^\s\])\"']+", re.I)


def extract_product_page_url(text: str, *, exclude: str = "") -> str:
    """从文本中提取非短视频平台的产品页链接。"""
    for m in GENERIC_URL_RE.finditer(text or ""):
        url = m.group(0).rstrip(".,;)")
        if exclude and url == exclude:
            continue
        if VIDEO_URL_RE.search(url):
            continue
        return url
    return ""
